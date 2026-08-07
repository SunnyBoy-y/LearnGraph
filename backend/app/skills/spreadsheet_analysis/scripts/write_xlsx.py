#!/usr/bin/env python3
"""Write a structured rows JSON into an .xlsx with a styled header (offline)."""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe(value: str) -> Path:
    path = Path(value)
    if path.is_absolute() or ".." in path.parts:
        raise RuntimeError("path must stay inside the sandbox workspace")
    return path


def _rows_from_input(src: Path) -> tuple[list[str], list[list]]:
    payload = json.loads(src.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "rows" in payload:
        rows = payload["rows"]
        columns = [str(c) for c in payload.get("columns", [])]
    elif isinstance(payload, list):
        rows = payload
        columns = []
    else:
        raise RuntimeError("input JSON must be a list of rows or {columns, rows}")
    if not isinstance(rows, list):
        raise RuntimeError("rows must be a list")
    if not columns and rows:
        if isinstance(rows[0], dict):
            columns = [str(k) for k in rows[0].keys()]
    if not columns:
        raise RuntimeError("could not infer columns")
    normalized: list[list] = []
    for row in rows:
        if isinstance(row, dict):
            normalized.append([row.get(c) for c in columns])
        elif isinstance(row, (list, tuple)):
            if len(row) != len(columns):
                raise RuntimeError("row length does not match columns")
            normalized.append(list(row))
        else:
            raise RuntimeError("row must be an object or array")
    return columns, normalized


def main() -> int:
    parser = argparse.ArgumentParser(description="Write structured rows to .xlsx (offline).")
    parser.add_argument("--input", required=True, help="workspace-relative rows.json path")
    parser.add_argument("--output", required=True, help="workspace-relative .xlsx path")
    parser.add_argument("--sheet", default="Sheet1")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    src = _safe(args.input)
    dst = _safe(args.output)
    if not src.is_file():
        raise RuntimeError(f"input file not found: {src}")
    if dst.exists() and not args.overwrite:
        raise RuntimeError("output already exists; pass --overwrite to replace")
    columns, rows = _rows_from_input(src)
    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet
    header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    for col in range(1, len(columns) + 1):
        width = max(8, min(60, max((len(str(columns[col - 1])) + 2, *(len(str(row[col - 1])) + 2 for row in rows[:200])))))
        ws.column_dimensions[get_column_letter(col)].width = width
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dst))
    data = dst.read_bytes()
    print(
        json.dumps(
            {
                "status": "ok",
                "input": str(src),
                "output": str(dst),
                "columns": len(columns),
                "rows": len(rows),
                "bytes": len(data),
                "sha256": hashlib.sha256(data).hexdigest(),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except RuntimeError as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(1) from exc
